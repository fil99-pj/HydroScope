# -*- coding: utf-8 -*-
"""
excel_import.py
================

Modulo di importazione dei dati di monitoraggio piezometrico da file
Excel (.xlsx) o CSV.

Il plugin supporta due modalita' di input, scelte automaticamente in base
alle colonne presenti nel file (e confermabili dall'utente in dialog.py):

1. MODALITA' "FILE UNICO"
   Il file contiene sia l'anagrafica dei piezometri (Nome, X, Y, Quota
   testa) sia i dati di campagna (Data, Soggiacenza, Spessore LNAPL),
   con piu' righe per piezometro (una per ogni data di campagna).
   Il modulo estrae da qui sia l'anagrafica (deduplicata per nome) sia
   le serie di campagna.

2. MODALITA' "SOLO CAMPAGNA"
   Il file contiene solo Nome piezometro, Data, Soggiacenza, Spessore
   LNAPL (nessuna colonna X/Y/Quota testa). In questo caso l'anagrafica
   deve provenire da un layer GIS piezometri gia' esistente in QGIS,
   fornito separatamente dall'utente nella UI, e il join viene fatto
   altrove (in mainPlugin.py / hydro_calculations.py) sul nome piezometro.

Il modulo NON dipende da PyQGIS: usa solo openpyxl e il modulo csv
standard, quindi e' testabile in isolamento.
"""

import os
import csv

from . import utils


# ---------------------------------------------------------------------------
# Mappatura nomi colonna flessibile (case-insensitive, sinonimi comuni)
# ---------------------------------------------------------------------------

# Ogni voce: nome_campo_interno -> lista di possibili intestazioni di colonna
# (confrontate ignorando maiuscole/minuscole e spazi superflui)
COLUMN_ALIASES = {
    "nome": ["nome", "nome piezometro", "id", "id piezometro", "piezometro", "codice"],
    "x": ["x", "coordinata x", "coord_x", "lon", "longitudine", "est", "easting"],
    "y": ["y", "coordinata y", "coord_y", "lat", "latitudine", "nord", "northing"],
    "quota_testa": [
        "quota testa", "quota testa piezometro", "quota_testa", "quota testa (m s.l.m.)",
        "quota testa pozzo", "z", "quota boccapozzo", "quota testa (m)",
    ],
    "quota_fondo_filtro": [
        "quota fondo filtro", "fondo filtro", "quota_fondo_filtro", "fondo filtro (m)",
    ],
    "data": ["data", "data rilievo", "data_rilievo", "data campagna", "data monitoraggio"],
    "soggiacenza": [
        "soggiacenza", "soggiacenza (m)", "soggiacenza_m", "profondita falda",
        "profondita' falda", "depth to water",
    ],
    "spessore_lnapl": [
        "spessore lnapl", "spessore_lnapl", "lnapl", "spessore lnapl (m)",
        "spessore lnapl (cm)", "spessore lnapl (mm)", "spessore prodotto",
        "spessore prodotto libero", "free product thickness",
    ],
}

# Colonne che richiedono di sapere l'unita' (m, cm oppure mm); si tenta di
# dedurla dall'intestazione stessa, altrimenti si usa il default passato
# dalla UI (vedi parametro lnapl_unit_default in read_campaign_file).
# NOTA: l'ordine di controllo conta - "mm" deve essere verificato PRIMA di
# "m" da solo, altrimenti l'hint "_m" matcherebbe anche dentro "mm".
LNAPL_UNIT_HINTS_MM = ["mm"]
LNAPL_UNIT_HINTS_CM = ["cm"]
LNAPL_UNIT_HINTS_M = ["(m)", " m)", "_m"]


class ImportError_(Exception):
    """Eccezione dedicata per errori di importazione (alias per evitare
    collisione con builtin ImportError)."""
    pass


def _normalize_header(header):
    return str(header).strip().lower()


def _match_column(headers_normalized, aliases):
    """Trova l'indice di colonna che corrisponde a uno degli alias dati.

    :param headers_normalized: lista di intestazioni normalizzate (lower/strip)
    :param aliases: lista di possibili nomi colonna per il campo cercato
    :return: indice colonna (int) oppure None se non trovata
    """
    for alias in aliases:
        alias_norm = alias.strip().lower()
        for idx, header in enumerate(headers_normalized):
            if header == alias_norm:
                return idx
    # fallback: match parziale (contiene)
    for alias in aliases:
        alias_norm = alias.strip().lower()
        for idx, header in enumerate(headers_normalized):
            if alias_norm in header:
                return idx
    return None


def _detect_lnapl_unit(header_text, default_unit="cm"):
    """Deduce l'unita' di misura del campo LNAPL dall'intestazione di colonna.

    :param header_text: testo intestazione originale (non normalizzato)
    :param default_unit: unita' da usare se non deducibile dall'header
    :return: "mm", "cm" oppure "m"
    """
    text = (header_text or "").lower()
    # "mm" va controllato PRIMA di "m": l'hint "_m"/" m)" altrimenti
    # matcherebbe erroneamente anche dentro "mm" (es. "spessore (mm)").
    for hint in LNAPL_UNIT_HINTS_MM:
        if hint in text:
            return "mm"
    for hint in LNAPL_UNIT_HINTS_CM:
        if hint in text:
            return "cm"
    for hint in LNAPL_UNIT_HINTS_M:
        if hint in text:
            return "m"
    return default_unit


def _read_rows_from_file(filepath):
    """Legge tutte le righe da un file .xlsx o .csv come liste di celle.

    :param filepath: percorso del file
    :return: tupla (headers, rows) dove headers e' una lista di stringhe
             e rows e' una lista di liste (una per riga di dati)
    :raises ImportError_: se il formato non e' supportato o il file e' vuoto
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:
            raise ImportError_(
                "La libreria 'openpyxl' non e' disponibile nell'ambiente Python "
                "di QGIS. Installarla con: pip install openpyxl (nel Python "
                "usato da QGIS)."
            )
        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        except Exception as exc:
            raise ImportError_("Impossibile aprire il file Excel: {0}".format(exc))

        sheet = workbook.active
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            raise ImportError_("Il file Excel e' vuoto.")

        headers = [str(c) if c is not None else "" for c in all_rows[0]]
        data_rows = [list(r) for r in all_rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        return headers, data_rows

    elif ext == ".csv":
        try:
            with open(filepath, "r", encoding="utf-8-sig", newline="") as fh:
                sample = fh.read(4096)
                fh.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
                except csv.Error:
                    dialect = csv.excel
                    dialect.delimiter = ";" if sample.count(";") >= sample.count(",") else ","
                reader = csv.reader(fh, dialect)
                all_rows = [row for row in reader]
        except Exception as exc:
            raise ImportError_("Impossibile aprire il file CSV: {0}".format(exc))

        if not all_rows:
            raise ImportError_("Il file CSV e' vuoto.")

        headers = all_rows[0]
        data_rows = [r for r in all_rows[1:] if any(c.strip() != "" for c in r)]
        return headers, data_rows

    else:
        raise ImportError_(
            "Formato file non supportato: '{0}'. Usare .xlsx o .csv.".format(ext)
        )


def detect_input_mode(filepath):
    """Determina automaticamente se il file e' in modalita' 'file unico'
    (contiene anagrafica completa) o 'solo campagna' (richiede layer GIS).

    :param filepath: percorso del file Excel/CSV
    :return: tupla (mode, headers) dove mode e' "full" oppure "campaign_only"
    """
    headers, _ = _read_rows_from_file(filepath)
    headers_norm = [_normalize_header(h) for h in headers]

    has_x = _match_column(headers_norm, COLUMN_ALIASES["x"]) is not None
    has_y = _match_column(headers_norm, COLUMN_ALIASES["y"]) is not None
    has_quota = _match_column(headers_norm, COLUMN_ALIASES["quota_testa"]) is not None

    if has_x and has_y and has_quota:
        return "full", headers
    return "campaign_only", headers


def read_monitoring_file(filepath, lnapl_unit_default="cm"):
    """Legge il file di monitoraggio (Excel o CSV) e restituisce i dati
    strutturati, in entrambe le modalita' supportate.

    Formato di ritorno (dict):
        {
            "mode": "full" | "campaign_only",
            "piezometers": {
                "PZ01": {"x": 123.4, "y": 456.7, "quota_testa": 10.5,
                         "quota_fondo_filtro": 5.0 (opzionale, None se assente)},
                ...
            },   # presente solo se mode == "full", altrimenti dict vuoto
            "campaigns": [
                {
                    "nome": "PZ01",
                    "data": datetime.date(...),
                    "soggiacenza": 3.2,                # metri
                    "spessore_lnapl_m": 0.08,           # metri (convertito)
                    "lnapl_class": "moderato",
                },
                ...
            ],
            "dates": [datetime.date(...), ...]   # date distinte ordinate
        }

    :param filepath: percorso del file
    :param lnapl_unit_default: unita' di default per lo spessore LNAPL se
        non deducibile dall'intestazione di colonna ("cm" o "m")
    :raises ImportError_: per errori di formato o colonne mancanti
    :raises utils.ValidationError: per righe con dati non validi
    """
    headers, rows = _read_rows_from_file(filepath)
    headers_norm = [_normalize_header(h) for h in headers]

    idx_nome = _match_column(headers_norm, COLUMN_ALIASES["nome"])
    idx_data = _match_column(headers_norm, COLUMN_ALIASES["data"])
    idx_sogg = _match_column(headers_norm, COLUMN_ALIASES["soggiacenza"])
    idx_lnapl = _match_column(headers_norm, COLUMN_ALIASES["spessore_lnapl"])

    idx_x = _match_column(headers_norm, COLUMN_ALIASES["x"])
    idx_y = _match_column(headers_norm, COLUMN_ALIASES["y"])
    idx_quota = _match_column(headers_norm, COLUMN_ALIASES["quota_testa"])
    idx_fondo = _match_column(headers_norm, COLUMN_ALIASES["quota_fondo_filtro"])

    required_missing = []
    if idx_nome is None:
        required_missing.append("Nome piezometro")
    if idx_data is None:
        required_missing.append("Data rilievo")
    if idx_sogg is None:
        required_missing.append("Soggiacenza")
    if required_missing:
        raise ImportError_(
            "Colonne obbligatorie non trovate nel file: {0}. "
            "Intestazioni rilevate: {1}".format(
                ", ".join(required_missing), ", ".join(h for h in headers if h)
            )
        )

    mode = "full" if (idx_x is not None and idx_y is not None and idx_quota is not None) else "campaign_only"

    lnapl_unit = _detect_lnapl_unit(
        headers[idx_lnapl] if idx_lnapl is not None else "", lnapl_unit_default
    )

    piezometers = {}
    campaigns = []
    dates_set = set()
    errors = []

    for row_number, row in enumerate(rows, start=2):  # start=2: riga 1 e' header
        def cell(idx):
            if idx is None or idx >= len(row):
                return None
            value = row[idx]
            if isinstance(value, str):
                value = value.strip()
                if value == "":
                    return None
            return value

        nome = cell(idx_nome)
        if nome in (None, ""):
            continue  # riga vuota o senza identificativo: saltata silenziosamente
        nome = str(nome).strip()

        raw_date = cell(idx_data)
        data_val = utils.parse_date(raw_date)
        if data_val is None:
            errors.append(
                "Riga {0} ({1}): data non valida o mancante ('{2}').".format(
                    row_number, nome, raw_date
                )
            )
            continue

        soggiacenza = utils.safe_float(cell(idx_sogg))
        if soggiacenza is None:
            errors.append(
                "Riga {0} ({1}): soggiacenza non valida o mancante.".format(row_number, nome)
            )
            continue

        lnapl_raw = utils.safe_float(cell(idx_lnapl)) if idx_lnapl is not None else None
        lnapl_m = utils.to_meters(lnapl_raw, lnapl_unit)
        lnapl_class = utils.classify_lnapl_thickness(lnapl_m)

        campaigns.append({
            "nome": nome,
            "data": data_val,
            "soggiacenza": soggiacenza,
            "spessore_lnapl_m": lnapl_m if lnapl_m is not None else 0.0,
            "lnapl_class": lnapl_class,
        })
        dates_set.add(data_val)

        if mode == "full":
            x = utils.safe_float(cell(idx_x))
            y = utils.safe_float(cell(idx_y))
            quota_testa = utils.safe_float(cell(idx_quota))
            quota_fondo = utils.safe_float(cell(idx_fondo)) if idx_fondo is not None else None

            if nome not in piezometers:
                if x is None or y is None or quota_testa is None:
                    errors.append(
                        "Riga {0} ({1}): anagrafica incompleta (X/Y/Quota testa) "
                        "alla prima occorrenza del piezometro.".format(row_number, nome)
                    )
                else:
                    piezometers[nome] = {
                        "x": x,
                        "y": y,
                        "quota_testa": quota_testa,
                        "quota_fondo_filtro": quota_fondo,
                    }

    if errors:
        utils.log(
            "Import completato con {0} righe scartate per errori di validazione."
            .format(len(errors)),
            level="WARNING",
        )
        for err in errors[:50]:  # limite per non intasare il log
            utils.log(err, level="WARNING")

    if not campaigns:
        raise ImportError_(
            "Nessun record valido trovato nel file. Verificare il contenuto e "
            "le intestazioni delle colonne."
        )

    return {
        "mode": mode,
        "piezometers": piezometers,
        "campaigns": campaigns,
        "dates": sorted(dates_set),
        "row_errors": errors,
    }


def write_example_files(output_dir):
    """Genera file Excel/CSV di esempio (modalita' completa e solo-campagna)
    nella cartella indicata. Utile per la documentazione e per i test.

    :param output_dir: cartella di destinazione (creata se non esiste)
    :return: lista dei percorsi file creati
    """
    os.makedirs(output_dir, exist_ok=True)
    created = []

    try:
        import openpyxl
    except ImportError:
        openpyxl = None

    # --- Esempio modalita' "full" --------------------------------------
    full_rows = [
        ["Nome piezometro", "X", "Y", "Quota testa (m s.l.m.)", "Quota fondo filtro",
         "Data rilievo", "Soggiacenza (m)", "Spessore LNAPL (cm)"],
        ["PZ01", 452310.5, 4983210.2, 102.35, 95.00, "15/03/2026", 3.20, 0],
        ["PZ02", 452365.1, 4983180.7, 101.90, 94.50, "15/03/2026", 2.95, 2.5],
        ["PZ03", 452410.8, 4983225.4, 102.10, 95.10, "15/03/2026", 3.05, 12.0],
        ["PZ04", 452290.3, 4983150.9, 101.60, 94.20, "15/03/2026", 2.80, 25.0],
        ["PZ05", 452350.0, 4983260.0, 102.50, 95.50, "15/03/2026", 3.40, 0],
        ["PZ01", 452310.5, 4983210.2, 102.35, 95.00, "10/06/2026", 3.55, 0],
        ["PZ02", 452365.1, 4983180.7, 101.90, 94.50, "10/06/2026", 3.10, 1.0],
        ["PZ03", 452410.8, 4983225.4, 102.10, 95.10, "10/06/2026", 3.20, 8.5],
        ["PZ04", 452290.3, 4983150.9, 101.60, 94.20, "10/06/2026", 2.95, 18.0],
        ["PZ05", 452350.0, 4983260.0, 102.50, 95.50, "10/06/2026", 3.60, 0],
    ]

    # --- Esempio modalita' "campaign_only" -----------------------------
    campaign_rows = [
        ["Nome piezometro", "Data rilievo", "Soggiacenza (m)", "Spessore LNAPL (cm)"],
        ["PZ01", "15/03/2026", 3.20, 0],
        ["PZ02", "15/03/2026", 2.95, 2.5],
        ["PZ03", "15/03/2026", 3.05, 12.0],
        ["PZ04", "15/03/2026", 2.80, 25.0],
        ["PZ05", "15/03/2026", 3.40, 0],
        ["PZ01", "10/06/2026", 3.55, 0],
        ["PZ02", "10/06/2026", 3.10, 1.0],
        ["PZ03", "10/06/2026", 3.20, 8.5],
        ["PZ04", "10/06/2026", 2.95, 18.0],
        ["PZ05", "10/06/2026", 3.60, 0],
    ]

    def _write_csv(path, rows):
        with open(path, "w", encoding="utf-8", newline="") as fh:
            writer = csv.writer(fh, delimiter=";")
            writer.writerows(rows)

    def _write_xlsx(path, rows, sheet_title):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title
        for row in rows:
            ws.append(row)
        wb.save(path)

    csv_full = os.path.join(output_dir, "esempio_campagna_completo.csv")
    _write_csv(csv_full, full_rows)
    created.append(csv_full)

    csv_campaign = os.path.join(output_dir, "esempio_solo_campagna.csv")
    _write_csv(csv_campaign, campaign_rows)
    created.append(csv_campaign)

    if openpyxl is not None:
        xlsx_full = os.path.join(output_dir, "esempio_campagna_completo.xlsx")
        _write_xlsx(xlsx_full, full_rows, "Campagna")
        created.append(xlsx_full)

        xlsx_campaign = os.path.join(output_dir, "esempio_solo_campagna.xlsx")
        _write_xlsx(xlsx_campaign, campaign_rows, "Campagna")
        created.append(xlsx_campaign)

    return created
